import datetime
import socket
import msvcrt as pause
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
import openpyxl as xl

host = "smtp.gmail.com"
port = 587

print("""
Warning: Please make sure to turn on the less secure apps so that this
         app can access the necessary permission from your G-mail account
         to send messages to the excel sheet users. Google: less-secure apps and you 
         may find that link of enabling less secure app access from the first link or
         so
""")
print("""
        Put that G-mail account and password from which you are
        sending all the message to the users

""")
# gmail_username = input("Put your G-mail (eg.abc@gmail.com): ")
# gmail_password = input("Put your G-mail password: ")
username = input("Enter your business gmail. (eg.yourname@gmail.com) \n Gmail Id >>> ")
password = input("Enter your gmail password \n Gmail Password >>> ")
from_email = username

#class for formatting keeping all the user details
class MessageUser():
    email_messages = []
    user_details = []
    messages = []
    base_message = """
    <html>
    <head>
    <title> Billing update </title>
    </head>
    <body style="background-color: orange; color: white">
    <h3><b>   <center> Hi, </b></h3><h1 style="color: red"> <center>   {name}  </h1>
<center>     All employee's salary has been sent to the employee's email on <h2> <b>{date}.</b> </h2><br/>
<center>    This message is sent to those employees who had given appropriate and their
<center>    personal active email to us. <br/>
<center>    Deducting all the taxes and according to the allowances your net salary is <h1 style="color: red">${total}</h1>
<center>    Have a great one!
Your employee id is <b style="color: #28968b">{eid}</b> <br/>
Team {companyname}
    </body>
    </html>

"""

    
#this adds user and its amount and stuffs to the user detail
    def add_user(self, name, amount, companyname=input("Enter your companyname >>> "), eid=None, email=None):
        name = name[0].upper() + name[1:].lower()
        amount = "%.2f" % (amount)
        detail = {
            "name": name,
            "amount": amount

        }
        today = datetime.date.today()
        time = datetime.datetime.now()
        text_ = '{time.hour}:{time.minute}:{time.second}'.format(time=time)
        text = '{today.month}/{today.day}/{today.year}'.format(today=today) + " at " + text_
        detail["date"] = text
        detail["companyname"] = companyname
        if eid is not None:
            detail["eid"] = eid
        if email is not None:
            detail["email"] = email
        self.user_details.append(detail)
# a function that returns all the appended details from user details
    def get_details(self):
        return self.user_details
#function for formatiing messages
    def make_messages(self):
        if len(self.user_details) > 0:
            for detail in self.get_details():
                name = detail["name"]
                amount = detail["amount"]
                date = detail["date"]
                eid = detail["eid"]
                companyname = detail["companyname"]
                message = self.base_message
                new_msg = message.format(
                    name=name,
                    date=date,
                    total=amount,
                    eid=eid,
                    companyname=companyname
                )
                user_email = detail.get("email")
                if user_email:
                    user_data = {
                        "email": user_email,
                        "message": new_msg
                    }
                    self.email_messages.append(user_data)
                else:
                    self.messages.append(new_msg)
            return self.messages
        return []
#send message using smtp module
    def send_email(self):
        self.make_messages()
        if len(self.email_messages) > 0:
            for detail in self.email_messages:
                user_email = detail["email"]
                user_message = detail["message"]
                # run email
                try:
                    email_conn = smtplib.SMTP(host, port)
                    email_conn.ehlo()
                    email_conn.starttls()
                    email_conn.login(username, password)
                    the_msg = MIMEMultipart("alternative")
                    the_msg['Subject'] = "Your monthly Salary"
                    the_msg["From"] = from_email
                    the_msg["To"] = user_email
                    part_1 = MIMEText(user_message, 'html')
                    the_msg.attach(part_1)
                    email_conn.sendmail(from_email, [user_email], the_msg.as_string())
                    email_conn.quit()
                except socket.gaierror:
                    print("Internet connection lost but the file is saved to 'excel_files' folder")
                except TypeError:
                    print("Email not send but the data are saved to 'excel_files' folder ")
                except smtplib.SMTPException:
                    print("error sending the message \n may be the user-email/password is incorrect")
            return True
        return False


# print("""
#
# Note: First make the primary file for salary sheet with no values just format and then
# create another excel file where you fill username and basic pay. that second file you create should be opened
# just put the name of the secondary file
# """)
wb = xl.load_workbook("payroll.xlsx") #load excel file
sheet = wb['Sheet1'] #There are many sheets in an excel file in this case we select the first sheet of excel file
for row in range(6, sheet.max_row + 1):
    names = sheet.cell(row, 3)
    if names.value is None:
        break
    basic_pay = sheet.cell(row, 4)
    if basic_pay.value is None:
        break
    medical_allowance = sheet.cell(row, 5)
    medical_allowance_percent = float(input(
        "Medical allowance percentage value without giving (%) for {name} \n amount(%): ".format(name=names.value)))
    medical_allowance.value = medical_allowance_percent / 100 * basic_pay.value
    house_rent_cell = sheet.cell(row, 6)
    house_rent_percent = float(
        input("House rent percent value without giving '%' sign for {name} \n percent: ".format(name=names.value)))
    house_rent_cell.value = house_rent_percent / 100 * basic_pay.value
    gross_pay_cell = sheet.cell(row, 7)
    gross_pay_cell.value = basic_pay.value + medical_allowance.value + house_rent_cell.value
    Tax = sheet.cell(row, 8)
    if gross_pay_cell.value > 15000:
        Tax.value = 15 / 100 * gross_pay_cell.value
    else:
        Tax.value = 13 / 100 * gross_pay_cell.value
    net_pay = sheet.cell(row, 9)
    net_pay.value = gross_pay_cell.value - Tax.value
    grade_cell = sheet.cell(row, 10)
    if net_pay.value > 15000:
        grade_cell.value = "Grade-I"
    else:
        grade_cell.value = "Grade-II"
    wb.save("payroll_corrected.xlsx")




wb_infos = xl.load_workbook("payroll_corrected.xlsx")
sheet_info = wb_infos["Sheet1"]
workbook = xl.load_workbook("userandamount.xlsx")
sheet_UandA = workbook["Sheet1"]
payroll_unf = xl.load_workbook("payroll.xlsx")
sheet_payroll = payroll_unf["Sheet1"]
obj = MessageUser()
i = 6
for rows in range(i, sheet_info.max_row + 1):  # payroll
    eid = sheet_UandA.cell(rows-2, 1)
    names = sheet_info.cell(rows, 3)
    print(names.value)
    if names.value is not None:
        eid.value = i-5
        i+=1
    username_ = sheet_UandA.cell(rows - 2, 2)
    print(username_.value)
    username_.value = names.value
    net_pay_ = sheet_info.cell(rows, 9)
    print(net_pay_.value)
    amounts = sheet_UandA.cell(rows - 2, 4)
    print(amounts.value)
    amounts.value = net_pay_.value
    emails = sheet_payroll.cell(rows, 11)
    emails_pcorrected = sheet_info.cell(rows, 11)
    emails_pcorrected.value = emails.value
    emailuanda = sheet_UandA.cell(rows-2, 5)
    emailuanda.value = emails_pcorrected.value
    if username_.value and amounts.value and emails.value is not None:
        obj.add_user(username_.value, amounts.value,eid=eid.value, email=emails.value)
    elif username_.value and amounts.value is not None:
        obj.add_user(username_.value, amounts.value, eid=eid.value)
    else:
        print("Nothing done correctly")

final_save = input("Put the name of the file to save it in the excel_files Folder : \n filename: ")
workbook.save(final_save + ".xlsx")

obj = MessageUser()
obj.make_messages()
for user_detail in obj.user_details:
    print(user_detail)



import os
if obj.send_email():
    print("emails sent successfully :D")
    pause.getch()

def correctsave():
        nameoffile = final_save + ".xlsx"
        new_dir = "excel_files/{saved_file}".format(saved_file=nameoffile)
        os.rename(nameoffile, new_dir)
        print("saved")


correctsave()
if not obj.send_email():
    correctsave()
    print("saved :) but emails not send")

obj.send_email()

